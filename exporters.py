"""
Modulo per l'esportazione dell'orario scolastico in formato Excel (.xlsx).
Include:
- Tabellone Generale Docenti (1 riga per docente con tutte le ore della settimana)
- Orario per singola Classe
- Orario per singolo Docente
- Orario per Aula (DADA)
- Report di Qualità e Statistiche
"""
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from models import TimetableProblem, DAYS_OF_WEEK
from solver import TimetableResult

def generate_excel_timetable(problem: TimetableProblem, result: TimetableResult) -> bytes:
    output = io.BytesIO()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    cfg = problem.config
    num_days = cfg.num_days
    days = cfg.active_days
    daily_hours = cfg.daily_hours[:num_days]
    max_hours = max(daily_hours)

    # Stili
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    day_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    day_font = Font(name="Calibri", size=11, bold=True, color="000000")
    sub_header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    sub_header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    
    lesson_font = Font(name="Calibri", size=10, bold=True)
    sub_font = Font(name="Calibri", size=9, italic=True, color="595959")
    
    free_day_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Verde chiaro
    gap_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")      # Arancio chiaro
    room_fill = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
    
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    thick_right = Border(
        right=Side(style='medium', color='1F4E78'),
        left=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # -------------------------------------------------------------
    # FOGLIO 1: TABELLONE GENERALE DOCENTI (RIGA PER DOCENTE)
    # -------------------------------------------------------------
    ws_tabellone = wb.create_sheet(title="Tabellone Generale Docenti")
    
    # Titolo Tabellone
    title_text = f"TABELLONE GENERALE ORARIO DOCENTI - {cfg.school_name.upper()}"
    ws_tabellone.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3 + num_days * max_hours)
    t_cell = ws_tabellone.cell(row=1, column=1, value=title_text)
    t_cell.font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    t_cell.alignment = Alignment(horizontal='left', vertical='center')

    # Riga 3: Header Giorni Raggruppati
    ws_tabellone.cell(row=3, column=1, value="Docente").fill = header_fill
    ws_tabellone.cell(row=3, column=1).font = header_font
    ws_tabellone.cell(row=3, column=1).alignment = center_align

    ws_tabellone.cell(row=3, column=2, value="Contratto").fill = header_fill
    ws_tabellone.cell(row=3, column=2).font = header_font
    ws_tabellone.cell(row=3, column=2).alignment = center_align

    ws_tabellone.cell(row=3, column=3, value="Tot Ore").fill = header_fill
    ws_tabellone.cell(row=3, column=3).font = header_font
    ws_tabellone.cell(row=3, column=3).alignment = center_align

    col_cursor = 4
    for d_idx, day_name in enumerate(days):
        h_in_day = daily_hours[d_idx]
        start_col = col_cursor
        end_col = col_cursor + h_in_day - 1
        
        ws_tabellone.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
        day_hdr_cell = ws_tabellone.cell(row=3, column=start_col, value=day_name.upper())
        day_hdr_cell.fill = header_fill
        day_hdr_cell.font = header_font
        day_hdr_cell.alignment = center_align
        
        col_cursor += h_in_day

    # Riga 4: Sub-Header delle Ore (1ª, 2ª, 3ª...)
    ws_tabellone.cell(row=4, column=1, value="").fill = sub_header_fill
    ws_tabellone.cell(row=4, column=2, value="").fill = sub_header_fill
    ws_tabellone.cell(row=4, column=3, value="").fill = sub_header_fill
    
    col_cursor = 4
    for d_idx in range(num_days):
        for h in range(daily_hours[d_idx]):
            h_cell = ws_tabellone.cell(row=4, column=col_cursor, value=f"{h+1}ª")
            h_cell.fill = sub_header_fill
            h_cell.font = sub_header_font
            h_cell.alignment = center_align
            h_cell.border = thin_border
            col_cursor += 1

    # Righe per ciascun Docente
    row_cursor = 5
    for t_id, teacher in problem.teachers.items():
        # Calcola totale ore assegnate
        t_assignments = [a for a in problem.assignments if a.teacher_id == t_id or t_id in a.co_teacher_ids]
        tot_hours = sum(a.hours_per_week for a in t_assignments)
        
        is_pt = getattr(teacher, "is_part_time", False)
        max_w = getattr(teacher, "max_working_days", None)
        contratto_txt = f"PT (max {max_w} gg)" if (is_pt and max_w) else ("Part-Time" if is_pt else "Tempo Pieno")

        # Col 1: Docente
        c_doc = ws_tabellone.cell(row=row_cursor, column=1, value=teacher.name)
        c_doc.font = Font(name="Calibri", size=10, bold=True)
        c_doc.border = thin_border
        
        # Col 2: Contratto
        c_cont = ws_tabellone.cell(row=row_cursor, column=2, value=contratto_txt)
        c_cont.font = Font(name="Calibri", size=9)
        c_cont.alignment = center_align
        c_cont.border = thin_border
        
        # Col 3: Totale Ore
        c_tot = ws_tabellone.cell(row=row_cursor, column=3, value=tot_hours)
        c_tot.font = Font(name="Calibri", size=10, bold=True)
        c_tot.alignment = center_align
        c_tot.border = thin_border

        # Calcola se il docente lavora in ciascun giorno
        day_has_lessons = [False] * num_days
        for d_idx in range(num_days):
            for h in range(daily_hours[d_idx]):
                if t_id in result.grid_by_teacher and result.grid_by_teacher[t_id][d_idx][h] is not None:
                    day_has_lessons[d_idx] = True
                    break

        col_cursor = 4
        for d_idx in range(num_days):
            is_day_free = not day_has_lessons[d_idx]
            
            # Calcola buchi nel giorno se presente
            first_l = None
            last_l = None
            if not is_day_free:
                lessons_in_day = [result.grid_by_teacher[t_id][d_idx][hh] is not None for hh in range(daily_hours[d_idx])]
                first_l = next((idx for idx, val in enumerate(lessons_in_day) if val), None)
                last_l = next((idx for idx in reversed(range(len(lessons_in_day))) if lessons_in_day[idx]), None)

            for h in range(daily_hours[d_idx]):
                cell = ws_tabellone.cell(row=row_cursor, column=col_cursor)
                cell.border = thin_border
                cell.alignment = center_align

                if is_day_free:
                    cell.value = "LIB"
                    cell.fill = free_day_fill
                    cell.font = Font(name="Calibri", size=8, bold=True, color="276A3C")
                else:
                    slot_info = result.grid_by_teacher.get(t_id, [])[d_idx][h] if t_id in result.grid_by_teacher else None
                    if slot_info:
                        # Mostra Classe, Materia e Aula/Palestra
                        cell_txt = f"{slot_info.class_name}\n({slot_info.subject_name[:5]})"
                        clean_r = slot_info.room_name.split("(")[0].strip().replace("ª", "") if getattr(slot_info, "room_name", None) else ""
                        if clean_r:
                            cell_txt += f"\n📍{clean_r[:7]}"
                        if getattr(slot_info, "is_compresenza", False) or getattr(slot_info, "compresenza_text", ""):
                            cell_txt += "\n👥"
                        cell.value = cell_txt
                        cell.font = Font(name="Calibri", size=8.5, bold=True)
                    else:
                        # Controlla se è ora buca
                        if first_l is not None and last_l is not None and first_l < h < last_l:
                            cell.value = "BUCA"
                            cell.fill = gap_fill
                            cell.font = Font(name="Calibri", size=8, bold=True, color="C00000")
                        else:
                            cell.value = "-"
                            cell.font = sub_font
                
                col_cursor += 1

        ws_tabellone.row_dimensions[row_cursor].height = 28
        row_cursor += 1

    # Larghezza colonne tabellone
    ws_tabellone.column_dimensions['A'].width = 28
    ws_tabellone.column_dimensions['B'].width = 16
    ws_tabellone.column_dimensions['C'].width = 10
    for c_i in range(4, col_cursor + 1):
        ws_tabellone.column_dimensions[get_column_letter(c_i)].width = 9

    # -------------------------------------------------------------
    # FOGLIO 2: ORARIO PER CLASSE
    # -------------------------------------------------------------
    ws_classes = wb.create_sheet(title="Orario Classi")
    current_row = 1

    for c_id, school_class in problem.classes.items():
        ws_classes.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_days + 1)
        title_text = f"ORARIO CLASSE: {school_class.name}"
        if cfg.is_dada:
            title_text += "  [Modello DADA - Aule Disciplinari]"
        title_cell = ws_classes.cell(row=current_row, column=1, value=title_text)
        title_cell.font = Font(name="Calibri", size=13, bold=True, color="1F4E78")
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        ws_classes.cell(row=current_row, column=1, value="Ora").fill = header_fill
        ws_classes.cell(row=current_row, column=1).font = header_font
        ws_classes.cell(row=current_row, column=1).alignment = center_align

        for d_idx, day_name in enumerate(days):
            col_idx = d_idx + 2
            cell = ws_classes.cell(row=current_row, column=col_idx, value=day_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        current_row += 1

        for h in range(max_hours):
            hour_label_cell = ws_classes.cell(row=current_row, column=1, value=f"{h+1}ª Ora")
            hour_label_cell.fill = day_fill
            hour_label_cell.font = day_font
            hour_label_cell.alignment = center_align
            hour_label_cell.border = thin_border

            for d_idx in range(num_days):
                col_idx = d_idx + 2
                cell = ws_classes.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_align

                if h < daily_hours[d_idx]:
                    slot_info = result.grid_by_class.get(c_id, [])[d_idx][h] if c_id in result.grid_by_class else None
                    if slot_info:
                        text = f"{slot_info.subject_name}\n({slot_info.teacher_name})"
                        if getattr(slot_info, "is_compresenza", False) or getattr(slot_info, "compresenza_text", ""):
                            c_t = getattr(slot_info, "compresenza_text", "") or "Compresenza"
                            text += f"\n👥 {c_t}"
                        if slot_info.room_name:
                            text += f"\n📍 {slot_info.room_name}"
                        cell.value = text
                        cell.font = lesson_font
                    else:
                        cell.value = "-"
                        cell.font = sub_font
                else:
                    cell.value = ""
            
            ws_classes.row_dimensions[current_row].height = 42
            current_row += 1

        current_row += 2

    # -------------------------------------------------------------
    # FOGLIO 3: ORARIO PER DOCENTE
    # -------------------------------------------------------------
    ws_teachers = wb.create_sheet(title="Orario Docenti (Dettagliato)")
    current_row = 1

    for t_id, teacher in problem.teachers.items():
        ws_teachers.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_days + 1)
        t_title = f"ORARIO: {teacher.name}"
        if teacher.free_day_1:
            t_title += f"  (Giorno Libero: {teacher.free_day_1})"
        title_cell = ws_teachers.cell(row=current_row, column=1, value=t_title)
        title_cell.font = Font(name="Calibri", size=13, bold=True, color="1F4E78")
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

        ws_teachers.cell(row=current_row, column=1, value="Ora").fill = header_fill
        ws_teachers.cell(row=current_row, column=1).font = header_font
        ws_teachers.cell(row=current_row, column=1).alignment = center_align

        for d_idx, day_name in enumerate(days):
            col_idx = d_idx + 2
            cell = ws_teachers.cell(row=current_row, column=col_idx, value=day_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
        current_row += 1

        day_has_lessons = [False] * num_days
        for d_idx in range(num_days):
            for h in range(daily_hours[d_idx]):
                if t_id in result.grid_by_teacher and result.grid_by_teacher[t_id][d_idx][h] is not None:
                    day_has_lessons[d_idx] = True
                    break

        for h in range(max_hours):
            hour_label_cell = ws_teachers.cell(row=current_row, column=1, value=f"{h+1}ª Ora")
            hour_label_cell.fill = day_fill
            hour_label_cell.font = day_font
            hour_label_cell.alignment = center_align
            hour_label_cell.border = thin_border

            for d_idx in range(num_days):
                col_idx = d_idx + 2
                cell = ws_teachers.cell(row=current_row, column=col_idx)
                cell.border = thin_border
                cell.alignment = center_align

                if h < daily_hours[d_idx]:
                    if not day_has_lessons[d_idx]:
                        cell.value = "LIBERO"
                        cell.fill = free_day_fill
                        cell.font = Font(name="Calibri", size=10, bold=True, color="276A3C")
                    else:
                        slot_info = result.grid_by_teacher.get(t_id, [])[d_idx][h] if t_id in result.grid_by_teacher else None
                        if slot_info:
                            text = f"Classe {slot_info.class_name}\n{slot_info.subject_name}"
                            if getattr(slot_info, "is_compresenza", False) or getattr(slot_info, "compresenza_text", ""):
                                c_t = getattr(slot_info, "compresenza_text", "") or "Compresenza"
                                text += f"\n👥 {c_t}"
                            if slot_info.room_name:
                                text += f"\n📍 {slot_info.room_name}"
                            cell.value = text
                            cell.font = lesson_font
                        else:
                            lessons_in_day = [result.grid_by_teacher[t_id][d_idx][hh] is not None for hh in range(daily_hours[d_idx])]
                            first_lesson = next((idx for idx, val in enumerate(lessons_in_day) if val), None)
                            last_lesson = next((idx for idx in reversed(range(len(lessons_in_day))) if lessons_in_day[idx]), None)
                            
                            if first_lesson is not None and last_lesson is not None and first_lesson < h < last_lesson:
                                cell.value = "BUCA"
                                cell.fill = gap_fill
                                cell.font = Font(name="Calibri", size=9, bold=True, color="C00000")
                            else:
                                cell.value = "-"
                                cell.font = sub_font
                else:
                    cell.value = ""

            ws_teachers.row_dimensions[current_row].height = 40
            current_row += 1

        current_row += 2

    # -------------------------------------------------------------
    # FOGLIO 4: OCCUPAZIONE AULE / DADA
    # -------------------------------------------------------------
    if problem.rooms:
        ws_rooms = wb.create_sheet(title="Orario Aule (DADA)")
        current_row = 1

        for r_id, room in problem.rooms.items():
            ws_rooms.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=num_days + 1)
            room_title = f"AULA: {room.name}"
            if room.subject_ids:
                subj_names = [problem.subjects[s].name for s in room.subject_ids if s in problem.subjects]
                room_title += f" (Discipline: {', '.join(subj_names)})"
            title_cell = ws_rooms.cell(row=current_row, column=1, value=room_title)
            title_cell.font = Font(name="Calibri", size=13, bold=True, color="1F4E78")
            title_cell.alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1

            ws_rooms.cell(row=current_row, column=1, value="Ora").fill = header_fill
            ws_rooms.cell(row=current_row, column=1).font = header_font
            ws_rooms.cell(row=current_row, column=1).alignment = center_align

            for d_idx, day_name in enumerate(days):
                col_idx = d_idx + 2
                cell = ws_rooms.cell(row=current_row, column=col_idx, value=day_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
            current_row += 1

            for h in range(max_hours):
                hour_label_cell = ws_rooms.cell(row=current_row, column=1, value=f"{h+1}ª Ora")
                hour_label_cell.fill = day_fill
                hour_label_cell.font = day_font
                hour_label_cell.alignment = center_align
                hour_label_cell.border = thin_border

                for d_idx in range(num_days):
                    col_idx = d_idx + 2
                    cell = ws_rooms.cell(row=current_row, column=col_idx)
                    cell.border = thin_border
                    cell.alignment = center_align

                    if h < daily_hours[d_idx]:
                        slot_info = result.grid_by_room.get(r_id, [])[d_idx][h] if r_id in result.grid_by_room else None
                        if slot_info:
                            text = f"Classe {slot_info.class_name}\n{slot_info.subject_name}\n({slot_info.teacher_name})"
                            if getattr(slot_info, "is_compresenza", False) or getattr(slot_info, "compresenza_text", ""):
                                c_t = getattr(slot_info, "compresenza_text", "") or "Compresenza"
                                text += f"\n👥 {c_t}"
                            cell.value = text
                            cell.font = lesson_font
                        else:
                            cell.value = "Disponibile"
                            cell.fill = room_fill
                            cell.font = sub_font
                    else:
                        cell.value = ""

                ws_rooms.row_dimensions[current_row].height = 42
                current_row += 1

            current_row += 2

    # -------------------------------------------------------------
    # FOGLIO 5: STATISTICHE & REPORT QUALITÀ
    # -------------------------------------------------------------
    ws_stats = wb.create_sheet(title="Report & Desiderata")
    ws_stats.cell(row=1, column=1, value="REPORT QUALITÀ ORARIO SCOLASTICO").font = Font(name="Calibri", size=14, bold=True, color="1F4E78")
    
    stats_data = [
        ("Modello Scolastico", "DADA (Ambienti di Apprendimento)" if cfg.is_dada else "Tradizionale"),
        ("Tempo di calcolo", f"{result.solve_time} secondi"),
        ("Stato Soluzione", "Ottimale" if result.status == "OPTIMAL" else "Valida / Ammissibile"),
        ("Giorni Liberi 1ª Scelta Soddisfatti", f"{result.free_days_satisfied_first} su {result.free_days_total_first}"),
        ("Giorni Liberi 2ª Scelta Soddisfatti", f"{result.free_days_satisfied_second} su {result.free_days_total_second}"),
        ("Ore Doppie Didattiche Soddisfatte", f"{result.double_hours_satisfied} su {result.double_hours_total}"),
        ("Totale Ore Buche (Docenti)", f"{result.total_gap_hours} ore complessive"),
        ("Ingressi Posticipati Concessi (No 1ª ora)", f"{result.late_entry_satisfied} su {result.late_entry_total}" if result.late_entry_total > 0 else "Nessuna richiesta"),
        ("Uscite Anticipate Concesse (No ult. ora)", f"{result.early_exit_satisfied} su {result.early_exit_total}" if result.early_exit_total > 0 else "Nessuna richiesta"),
        ("Slot Sconsigliati Evitati con Successo", f"{result.soft_slots_satisfied} su {result.soft_slots_total}" if result.soft_slots_total > 0 else "Nessuna richiesta"),
    ]
    
    for r_idx, (label, val) in enumerate(stats_data, start=3):
        ws_stats.cell(row=r_idx, column=1, value=label).font = Font(name="Calibri", size=11, bold=True)
        ws_stats.cell(row=r_idx, column=2, value=val).font = Font(name="Calibri", size=11)
        ws_stats.cell(row=r_idx, column=1).border = thin_border
        ws_stats.cell(row=r_idx, column=2).border = thin_border

    # Larghezza Colonne per i fogli classi e docenti
    for ws in [ws_classes, ws_teachers]:
        ws.column_dimensions['A'].width = 16
        for col_i in range(2, num_days + 3):
            ws.column_dimensions[get_column_letter(col_i)].width = 25

    if problem.rooms and "Orario Aule (DADA)" in wb.sheetnames:
        ws_r = wb["Orario Aule (DADA)"]
        ws_r.column_dimensions['A'].width = 16
        for col_i in range(2, num_days + 3):
            ws_r.column_dimensions[get_column_letter(col_i)].width = 25

    ws_stats.column_dimensions['A'].width = 38
    ws_stats.column_dimensions['B'].width = 28

    wb.save(output)
    return output.getvalue()
