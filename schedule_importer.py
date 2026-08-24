"""Modulo per l'importazione e il parsing di orari curricolari da fogli di calcolo Excel."""
import io
import openpyxl
from typing import Dict, List, Optional, Tuple, Any
from models import TimetableProblem, SchoolClass, Teacher, Subject, Classroom
from solver import TimetableResult, LessonSlotInfo

SlotInfo = LessonSlotInfo
Class = SchoolClass
Room = Classroom
from schedule_validator import validate_timetable, ValidationReport
from manual_editor_engine import _rebuild_result_views

def import_timetable_from_excel(
    file_bytes: bytes,
    problem: TimetableProblem
) -> Tuple[Optional[TimetableResult], ValidationReport, List[str]]:
    """Carica un file Excel con gli orari delle classi e genera un TimetableResult con audit di validita."""
    logs: List[str] = []
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    
    days = getattr(problem.config, "active_days", None) or getattr(problem.config, "days", ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì"])[:problem.config.num_days]
    daily_h = getattr(problem.config, "daily_hours", [6]*5)[:problem.config.num_days]
    
    # Mappe di lookup per riconoscimento intelligente (Case-Insensitive e strip)
    classes_by_name = {c.name.strip().lower(): c for c in problem.classes.values()}
    subjects_by_name = {s.name.strip().lower(): s for s in problem.subjects.values()}
    subjects_by_id = {s_id.strip().lower(): s for s_id, s in problem.subjects.items()}
    teachers_by_name = {t.name.strip().lower(): t for t in problem.teachers.values()}
    rooms_by_name = {r.name.strip().lower(): r for r in problem.rooms.values()}
    
    grid_by_class: Dict[str, List[List[Optional[SlotInfo]]]] = {
        c_id: [[None for _ in range(daily_h[d])] for d in range(len(days))]
        for c_id in problem.classes.keys()
    }
    
    # 1. STRATEGIA A: Un foglio per ciascuna classe (es. "1A", "Classe 1A", ecc.)
    matched_class_sheets = 0
    for sheet_name in wb.sheetnames:
        clean_sname = sheet_name.strip().lower().replace("classe", "").replace("cl.", "").strip()
        target_class: Optional[Class] = None
        
        for c_norm, c_obj in classes_by_name.items():
            clean_cname = c_norm.replace("classe", "").replace("cl.", "").strip()
            if clean_sname == clean_cname or sheet_name.strip().lower() == c_norm:
                target_class = c_obj
                break
                
        if not target_class or sheet_name.lower().startswith("orario"):
            continue
            
        matched_class_sheets += 1
        ws = wb[sheet_name]
        logs.append(f"Trovato foglio orario per classe: **{target_class.name}**")
        
        # Mappatura delle colonne dei giorni
        day_cols: Dict[int, int] = {} # d_idx -> col_idx (1-based)
        hour_rows: Dict[int, int] = {} # h_idx -> row_idx (1-based)
        
        # Trova intestazioni giorni nelle prime 5 righe
        for r in range(1, min(10, ws.max_row + 1)):
            for c in range(1, ws.max_column + 1):
                cell_val = str(ws.cell(row=r, column=c).value or "").strip().lower()
                for d_idx, d_name in enumerate(days):
                    if d_name.lower() in cell_val:
                        day_cols[d_idx] = c
                        
        # Trova righe delle ore nella prima colonna utile
        for r in range(1, ws.max_row + 1):
            for c in range(1, 4):
                cell_val = str(ws.cell(row=r, column=c).value or "").strip().lower()
                for h in range(max(daily_h)):
                    if f"{h+1}" in cell_val and ("ora" in cell_val or "h" in cell_val or cell_val.startswith(str(h+1))):
                        if h not in hour_rows:
                            hour_rows[h] = r

        if len(day_cols) < len(days):
            day_cols = {d: d + 2 for d in range(len(days))}
        if len(hour_rows) < max(daily_h):
            hour_rows = {h: h + 3 for h in range(max(daily_h))}
            
        for d in range(len(days)):
            col_idx = day_cols.get(d)
            if not col_idx:
                continue
            for h in range(daily_h[d]):
                row_idx = hour_rows.get(h)
                if not row_idx:
                    continue
                
                raw_text = str(ws.cell(row=row_idx, column=col_idx).value or "").strip()
                if not raw_text or raw_text in ("-", "None", "LIB"):
                    continue
                    
                slot = _parse_cell_to_slot(raw_text, target_class.id, problem, subjects_by_name, subjects_by_id, teachers_by_name, rooms_by_name)
                if slot:
                    grid_by_class[target_class.id][d][h] = slot

    # 2. STRATEGIA B: Foglio cumulativo "Orario Classi" con blocchi verticali
    if matched_class_sheets == 0:
        target_sheet = None
        for sname in wb.sheetnames:
            if "classi" in sname.lower() or "orario classi" in sname.lower():
                target_sheet = wb[sname]
                break
        if not target_sheet and len(wb.sheetnames) > 0:
            target_sheet = wb.active
            
        if target_sheet:
            logs.append(f"Lettura da foglio cumulativo classi: **{target_sheet.title}**")
            ws = target_sheet
            
            # Scansiona le righe per trovare intestazioni "ORARIO CLASSE: 1A"
            r = 1
            while r <= ws.max_row:
                cell_val = str(ws.cell(row=r, column=1).value or "").strip()
                found_c: Optional[SchoolClass] = None
                
                if "orario classe:" in cell_val.lower() or "classe" in cell_val.lower():
                    clean_header = cell_val.lower().replace("orario classe:", "").replace("classe", "").replace("ª", "").replace(" ", "").split("[")[0].strip()
                    for c_norm, c_obj in classes_by_name.items():
                        clean_cnorm = c_norm.replace("ª", "").replace(" ", "").lower()
                        if clean_cnorm == clean_header or clean_cnorm in clean_header or c_obj.id.lower() == clean_header:
                            found_c = c_obj
                            break
                            
                if found_c:
                    logs.append(f"Identificato blocco per classe **{found_c.name}** a riga {r}")
                    # Riga successiva: intestazioni giorni (Lunedì, Martedì...)
                    # Righe r+2 a r+2+max_h: ore di lezione
                    for h in range(max(daily_h)):
                        row_h = r + 2 + h
                        for d in range(len(days)):
                            col_d = d + 2
                            if h < daily_h[d] and row_h <= ws.max_row:
                                c_text = str(ws.cell(row=row_h, column=col_d).value or "").strip()
                                if c_text and c_text not in ("-", "None", "LIB"):
                                    slot = _parse_cell_to_slot(c_text, found_c.id, problem, subjects_by_name, subjects_by_id, teachers_by_name, rooms_by_name)
                                    if slot:
                                        grid_by_class[found_c.id][d][h] = slot
                    r += (max(daily_h) + 2)
                else:
                    r += 1

    # Costruisci il TimetableResult
    imported_result = TimetableResult(
        status="IMPORTED",
        solve_time=0.0,
        objective_value=0.0,
        grid_by_class=grid_by_class,
        grid_by_teacher={},
        grid_by_room={},
        gaps_by_teacher={}
    )
    
    # Rigenera viste docente, aule e statistiche
    _rebuild_result_views(problem, imported_result)
    
    # Audit e Validazione Conflitti
    report = validate_timetable(problem, imported_result)
    
    return imported_result, report, logs

def _parse_cell_to_slot(
    text: str,
    class_id: str,
    problem: TimetableProblem,
    subs_by_name: Dict[str, Subject],
    subs_by_id: Dict[str, Subject],
    teachers_by_name: Dict[str, Teacher],
    rooms_by_name: Dict[str, Room]
) -> Optional[SlotInfo]:
    """Interpreta una stringa cella ed estrae materia, docente e aula associati."""
    clean_lines = [l.strip() for l in text.split("\n") if l.strip()]
    if not clean_lines:
        clean_lines = [text.strip()]
        
    found_sub: Optional[Subject] = None
    found_t: Optional[Teacher] = None
    found_r: Optional[Room] = None
    
    # 1. Ricerca Materia (ordinamento per lunghezza decrescente per evitare match parziali)
    sorted_subs_by_name = sorted(subs_by_name.items(), key=lambda x: -len(x[0]))
    for line in clean_lines:
        norm_l = line.lower().replace("📖", "").replace("classe", "").strip()
        first_word = norm_l.split("(")[0].strip()
        for s_name_norm, s_obj in sorted_subs_by_name:
            s_base_norm = s_name_norm.split("(")[0].strip()
            if s_name_norm in norm_l or norm_l == s_name_norm or (len(s_base_norm) >= 4 and s_base_norm in norm_l):
                found_sub = s_obj
                break
        if not found_sub:
            for s_id_norm, s_obj in subs_by_id.items():
                if s_id_norm == norm_l or s_id_norm in norm_l.split():
                    found_sub = s_obj
                    break
        if found_sub:
            break
            
    # Se ancora non trovata, cerca tra le cattedre della classe
    class_assigns = [a for a in problem.assignments if a.class_id == class_id]
    if not found_sub:
        for a in sorted(class_assigns, key=lambda x: -len(problem.subjects[x.subject_id].name if x.subject_id in problem.subjects else "")):
            s_obj = problem.subjects.get(a.subject_id)
            if s_obj and (s_obj.name.lower() in text.lower() or s_obj.id.lower() in text.lower()):
                found_sub = s_obj
                found_t = problem.teachers.get(a.teacher_id)
                break

    if not found_sub:
        return None
        
    # 2. Ricerca Docente
    if not found_t:
        for line in clean_lines:
            norm_l = line.lower().replace("👤", "").replace("prof.", "").replace("prof.ssa", "").strip()
            for t_name_norm, t_obj in teachers_by_name.items():
                clean_tn = t_name_norm.replace("prof.", "").replace("prof.ssa", "").strip()
                if clean_tn in norm_l or norm_l in clean_tn:
                    found_t = t_obj
                    break
            if found_t:
                break
                
    if not found_t:
        for a in class_assigns:
            if a.subject_id == found_sub.id:
                found_t = problem.teachers.get(a.teacher_id)
                break

    # 3. Ricerca Aula
    for line in clean_lines:
        norm_l = line.lower().replace("📍", "").replace("aula", "").replace("lab.", "").strip()
        for r_name_norm, r_obj in rooms_by_name.items():
            if r_name_norm in norm_l or norm_l in r_name_norm:
                found_r = r_obj
                break
        if found_r:
            break
            
    if not found_r:
        if problem.config.is_dada and getattr(found_sub, "default_room_id", None):
            found_r = problem.rooms.get(found_sub.default_room_id)
        else:
            c_obj = problem.classes.get(class_id)
            fix_r_id = getattr(c_obj, "fixed_room_id", None)
            if fix_r_id:
                found_r = problem.rooms.get(fix_r_id)
            elif class_id in problem.rooms:
                found_r = problem.rooms[class_id]

    assign_id = ""
    for a in class_assigns:
        if a.subject_id == found_sub.id and (not found_t or a.teacher_id == found_t.id):
            assign_id = a.id
            break

    return SlotInfo(
        assignment_id=assign_id,
        class_id=class_id,
        class_name=problem.classes[class_id].name if class_id in problem.classes else class_id,
        subject_id=found_sub.id,
        subject_name=found_sub.name,
        subject_color=found_sub.color,
        teacher_id=found_t.id if found_t else "",
        teacher_name=found_t.name if found_t else "Docente Non Assegnato",
        room_id=found_r.id if found_r else "",
        room_name=found_r.name if found_r else ""
    )
