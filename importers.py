"""
Modulo per l'importazione ed esportazione di template CSV ed Excel (.xlsx).
Permette agli utenti di scaricare un template vuoto o di esempio, compilarlo con Excel o Calc,
e ricaricarlo per popolare automaticamente docenti, desiderata, classi, materie, cattedre e vincoli didattici.
"""
import io
import csv
import re
import pandas as pd
from typing import Tuple, List, Dict, Optional, Any
from models import (
    SchoolConfig, Teacher, SchoolClass, Subject, Classroom, 
    TeachingAssignment, TimetableProblem, DAYS_OF_WEEK,
    StudentDVA, SupportAssignment, EnhancementAssignment, ParallelGroup
)

# Colonne ordinate logicamente: prima i dati della Cattedra, poi i Desiderata del Docente
CSV_HEADERS = [
    "Docente",
    "CdC",
    "Classe",
    "Materia",
    "Ore_Settimanali",
    "Ore_Doppie",
    "Max_Ore_Giorno_Materia",
    "Part_Time",
    "Ore_Contratto",
    "Max_Giorni_Presenza",
    "Giorni_Liberi",
    "Entra_Tardi",
    "Esce_Presto",
    "Max_Ore_Giorno_Docente",
    "Max_Ore_Buche",
    "Slot_Sconsigliati",
    "Slot_Indisponibili"
]

# Mappatura colori predefiniti per le materie
DEFAULT_SUBJECT_COLORS = {
    "ita": "#e74c3c", "italiano": "#e74c3c", "lettere": "#e74c3c",
    "sto": "#e67e22", "storia": "#e67e22",
    "geo": "#d35400", "geografia": "#d35400",
    "mat": "#2980b9", "matematica": "#2980b9",
    "sci": "#27ae60", "scienze": "#27ae60",
    "ing": "#8e44ad", "inglese": "#8e44ad",
    "spa": "#9b59b6", "spagnolo": "#9b59b6", "seconda lingua": "#9b59b6", "francese": "#9b59b6", "tedesco": "#9b59b6",
    "tec": "#16a085", "tecnologia": "#16a085", "coding": "#16a085", "informatica": "#16a085",
    "mus": "#f39c12", "musica": "#f39c12", "strumento": "#f39c12",
    "art": "#e84393", "arte": "#e84393", "arte e immagine": "#e84393",
    "mot": "#00b894", "scienze motorie": "#00b894", "motoria": "#00b894", "ed. fisica": "#00b894",
    "rel": "#7f8c8d", "religione": "#7f8c8d", "materia alternativa": "#7f8c8d",
    "teatro": "#fd79a8", "laboratorio teatro": "#fd79a8", "laboratorio di teatro": "#fd79a8"
}

def _build_rows_data(sample_problem: Optional[TimetableProblem] = None) -> List[List[Any]]:
    """Costruisce le righe dati ordinate logicamente per CSV ed Excel."""
    rows = []
    if sample_problem and sample_problem.assignments:
        # Ordina per docente e classe per massima leggibilità
        sorted_assigns = sorted(
            sample_problem.assignments, 
            key=lambda a: (
                sample_problem.teachers.get(a.teacher_id).name if a.teacher_id in sample_problem.teachers else "",
                sample_problem.classes.get(a.class_id).name if a.class_id in sample_problem.classes else ""
            )
        )
        for a in sorted_assigns:
            t = sample_problem.teachers.get(a.teacher_id)
            c = sample_problem.classes.get(a.class_id)
            s = sample_problem.subjects.get(a.subject_id)
            
            t_name = t.name if t else a.teacher_id
            t_cdc = getattr(t, "cdc", "") if t else (getattr(s, "cdc", "") if s else "")
            c_name = c.name if c else a.class_id
            s_name = s.name if s else a.subject_id
            a_hours = a.hours_per_week
            a_double = "Si" if a.force_double_hours else "No"
            a_max_d = a.max_daily_hours or 2
            
            t_pt = "Si" if (t and getattr(t, "is_part_time", False)) else "No"
            t_ch = getattr(t, "contract_hours", 18) if t else 18
            t_mwd = getattr(t, "max_working_days", 5) if t else 5
            
            f_list = getattr(t, "free_days", []) if t else []
            if not f_list and t:
                if getattr(t, "free_day_1", None): f_list.append(t.free_day_1)
                if getattr(t, "free_day_2", None): f_list.append(t.free_day_2)
            giorni_liberi_str = ", ".join(f_list) if f_list else ""
            
            t_late = "Si" if (t and getattr(t, "prefer_late_entry", False)) else "No"
            t_early = "Si" if (t and getattr(t, "prefer_early_exit", False)) else "No"
            t_mdh = getattr(t, "max_daily_hours", 5) if t else 5
            t_mgh = getattr(t, "max_gap_hours", 2) if t else 2
            
            soft_str = ""
            if t and getattr(t, "soft_avoid_slots", []):
                slot_strs = []
                for d_i, h_i in t.soft_avoid_slots:
                    if d_i < len(DAYS_OF_WEEK):
                        slot_strs.append(f"{DAYS_OF_WEEK[d_i]} {h_i+1}")
                soft_str = ", ".join(slot_strs)
                
            unavail_str = ""
            if t and getattr(t, "unavailable_slots", []):
                slot_strs = []
                for d_i, h_i in t.unavailable_slots:
                    if d_i < len(DAYS_OF_WEEK):
                        slot_strs.append(f"{DAYS_OF_WEEK[d_i]} {h_i+1}")
                unavail_str = ", ".join(slot_strs)

            rows.append([
                t_name,
                t_cdc,
                c_name,
                s_name,
                a_hours,
                a_double,
                a_max_d,
                t_pt,
                t_ch or 18,
                t_mwd or 5,
                giorni_liberi_str,
                t_late,
                t_early,
                t_mdh,
                t_mgh,
                soft_str,
                unavail_str
            ])
    else:
        # Righe di esempio chiare ed esplicative per il template vuoto
        rows = [
            [
                "Prof.ssa Bianchi", "A-22", "1ª A", "Italiano", 6, "Si", 2,
                "No", 18, 5, "", "No", "Si", 5, 2, "", ""
            ],
            [
                "Prof.ssa Bianchi", "A-22", "1ª A", "Storia", 2, "No", 2,
                "No", 18, 5, "", "No", "Si", 5, 2, "", ""
            ],
            [
                "Prof. Verdi", "A-28", "1ª A", "Matematica", 4, "Si", 2,
                "No", 18, 5, "", "Si", "No", 5, 1, "", ""
            ],
            [
                "Prof. Verdi", "A-28", "1ª A", "Scienze", 2, "No", 2,
                "No", 18, 5, "", "Si", "No", 5, 1, "", ""
            ],
            [
                "Prof.ssa Colombo", "A-24", "1ª A", "Seconda Lingua (Spagnolo)", 2, "No", 2,
                "Si", 12, 3, "Lunedì, Mercoledì", "No", "No", 4, 1, "Mercoledì 1", "Lunedì 6"
            ]
        ]
    return rows

def generate_csv_template(sample_problem: Optional[TimetableProblem] = None) -> str:
    """
    Genera il contenuto CSV (con separatore punto e virgola ';' ideale per Excel italiano).
    """
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(CSV_HEADERS)
    for r in _build_rows_data(sample_problem):
        writer.writerow(r)
    return output.getvalue()

def generate_excel_template(sample_problem: Optional[TimetableProblem] = None) -> bytes:
    """
    Genera un file Excel (.xlsx) formattato con intestazioni in evidenza e larghezza colonne automatica.
    """
    rows = _build_rows_data(sample_problem)
    df = pd.DataFrame(rows, columns=CSV_HEADERS)
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Cattedre_e_Docenti', index=False)
        
        # Formattazione grafica del foglio Excel
        ws = writer.sheets['Cattedre_e_Docenti']
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        for col_num, col_name in enumerate(CSV_HEADERS, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            
            # Adatta la larghezza della colonna
            max_len = max([len(str(r[col_num-1])) for r in rows] + [len(col_name)]) + 4
            col_letter = cell.column_letter
            ws.column_dimensions[col_letter].width = max(max_len, 14)
            
        # Applica bordi a tutte le celle
        for row in ws.iter_rows(min_row=2, max_row=len(rows)+1, min_col=1, max_col=len(CSV_HEADERS)):
            for cell in row:
                cell.border = thin_border
                if cell.column in [5, 6, 7, 8, 9, 10, 12, 13, 14, 15]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                    
    return output.getvalue()


def _parse_bool(val: Any) -> bool:
    if not val or pd.isna(val):
        return False
    v = str(val).strip().lower()
    return v in ["si", "sì", "yes", "true", "1", "s", "y"]

def _parse_int(val: Any, default: int = 0) -> int:
    if val is None or pd.isna(val) or val == "":
        return default
    try:
        return int(float(str(val).strip().replace(",", ".")))
    except:
        return default

def _clean_id(name: str) -> str:
    return re.sub(r'[^a-zA-Z0-9_]', '_', name.strip().lower()).strip('_')

def _parse_slots_str(text: Any) -> List[List[int]]:
    """Converte stringhe come 'Mercoledì 1, Venerdì 5' o 'Lun 2; Ven 6' in [[day, hour], ...]"""
    if not text or pd.isna(text) or not str(text).strip():
        return []
    
    day_map = {
        "lun": 0, "luned": 0, "mon": 0,
        "mar": 1, "marted": 1, "tue": 1,
        "mer": 2, "mercol": 2, "wed": 2,
        "gio": 3, "gioved": 3, "thu": 3,
        "ven": 4, "venerd": 4, "fri": 4,
        "sab": 5, "sabat": 5, "sat": 5,
    }
    
    results = []
    parts = re.split(r'[,;/|]+', str(text))
    for part in parts:
        p = part.strip().lower()
        if not p:
            continue
        
        found_day = None
        for k, d_idx in day_map.items():
            if k in p:
                found_day = d_idx
                break
                
        num_match = re.search(r'(\d+)', p)
        if found_day is not None and num_match:
            h_num = int(num_match.group(1)) - 1 # 0-indexed
            if 0 <= h_num <= 8:
                results.append([found_day, h_num])
                
    return results

def _parse_free_days(text: Any) -> List[str]:
    """Estrae i giorni liberi da testo come 'Lunedì, Mercoledì' o 'Martedi'"""
    if not text or pd.isna(text) or not str(text).strip():
        return []
    
    day_patterns = {
        "Lunedì": ["lun"],
        "Martedì": ["mar"],
        "Mercoledì": ["mer"],
        "Giovedì": ["gio"],
        "Venerdì": ["ven"],
        "Sabato": ["sab"]
    }
    
    found = []
    parts = re.split(r'[,;/|]+', str(text).strip().lower())
    for part in parts:
        p = part.strip()
        for formal_name, subpatterns in day_patterns.items():
            if any(sub in p for sub in subpatterns):
                if formal_name not in found:
                    found.append(formal_name)
    return found

def parse_timetable_dataframe(df: pd.DataFrame, base_config: Optional[SchoolConfig] = None) -> Tuple[TimetableProblem, List[str]]:
    """
    Parser unificato e flessibile per DataFrame pandas (da CSV o Excel).
    Riconosce qualsiasi ordine delle colonne e molteplici varianti dei nomi di intestazione.
    """
    logs = []
    
    # Normalizza i nomi delle colonne
    clean_cols = [str(c).strip().lower().replace(" ", "_").replace("(", "").replace(")", "").replace("/", "_") for c in df.columns]
    
    def get_col_name(candidates: List[str]) -> Optional[str]:
        for c in candidates:
            for idx, col_clean in enumerate(clean_cols):
                if c == col_clean or c in col_clean:
                    return df.columns[idx]
        return None

    c_doc = get_col_name(["docente", "nome_docente", "insegnante", "prof"])
    c_cdc = get_col_name(["cdc", "classe_di_concorso", "concorso"])
    c_cls = get_col_name(["classe", "sezione", "classe_id"])
    c_sub = get_col_name(["materia", "disciplina", "insegnamento"])
    c_hrs = get_col_name(["ore_settimanali", "ore_settimana", "ore_sett", "ore", "monte_ore_materia"])
    c_dbl = get_col_name(["ore_doppie", "doppie", "blocco_2h", "blocco"])
    c_mda = get_col_name(["max_ore_giorno_materia", "max_giornaliere_materia", "max_ore_materia"])
    
    c_pt  = get_col_name(["part_time", "pt", "orario_ridotto"])
    c_ch  = get_col_name(["ore_contratto", "contratto", "monte_ore_docente", "ore_docente"])
    c_mwd = get_col_name(["max_giorni_presenza", "max_giorni", "giorni_presenza", "max_gg"])
    c_fds = get_col_name(["giorni_liberi", "giorno_libero", "liberi", "giorno_libero_1", "libero_1"])
    c_fd2 = get_col_name(["giorno_libero_2", "libero_2"])
    c_late = get_col_name(["entra_tardi", "tardi", "no_1a_ora", "no_1"])
    c_early = get_col_name(["esce_presto", "presto", "no_ultima_ora", "no_ult"])
    c_mdh = get_col_name(["max_ore_giorno_docente", "max_ore_giorno", "max_ore_docente"])
    c_mgh = get_col_name(["max_ore_buche", "buche", "max_gap"])
    c_sft = get_col_name(["slot_sconsigliati", "sconsigliati", "desiderata"])
    c_unv = get_col_name(["slot_indisponibili", "indisponibili", "indisponibilita"])

    if not c_doc or not c_cls or not c_sub:
        raise ValueError("Il file deve contenere almeno le colonne: 'Docente', 'Classe' e 'Materia'.")

    config = base_config or SchoolConfig(
        num_days=5,
        daily_hours=[6, 6, 6, 6, 6],
        school_name="Scuola Secondaria di I Grado",
        school_type="Secondaria I Grado (Scuola Media)"
    )

    teachers: Dict[str, Teacher] = {}
    classes: Dict[str, SchoolClass] = {}
    subjects: Dict[str, Subject] = {}
    assignments: List[TeachingAssignment] = []

    for _, row in df.iterrows():
        doc_val = str(row[c_doc]).strip() if pd.notna(row[c_doc]) else ""
        cls_val = str(row[c_cls]).strip() if pd.notna(row[c_cls]) else ""
        sub_val = str(row[c_sub]).strip() if pd.notna(row[c_sub]) else ""
        
        if not doc_val or not cls_val or not sub_val or doc_val.lower() == "nan":
            continue

        # 1. Parsing Docente
        t_id = "doc_" + _clean_id(doc_val)
        cdc_val = str(row[c_cdc]).strip() if (c_cdc and pd.notna(row[c_cdc])) else ""
        if cdc_val.lower() == "nan": cdc_val = ""
        
        if t_id not in teachers:
            is_pt = _parse_bool(row[c_pt]) if c_pt else False
            c_hours = _parse_int(row[c_ch], default=(12 if is_pt else 18)) if c_ch else (12 if is_pt else 18)
            m_days = _parse_int(row[c_mwd], default=(3 if is_pt else config.num_days)) if c_mwd else (3 if is_pt else config.num_days)
            
            f_days = []
            if c_fds and pd.notna(row[c_fds]):
                f_days.extend(_parse_free_days(row[c_fds]))
            if c_fd2 and pd.notna(row[c_fd2]):
                for fd in _parse_free_days(row[c_fd2]):
                    if fd not in f_days: f_days.append(fd)
                    
            late_val = _parse_bool(row[c_late]) if c_late else False
            early_val = _parse_bool(row[c_early]) if c_early else False
            mdh_val = _parse_int(row[c_mdh], default=5) if c_mdh else 5
            mgh_val = _parse_int(row[c_mgh], default=2) if c_mgh else 2
            
            soft_slots = _parse_slots_str(row[c_sft]) if c_sft else []
            unavail_slots = _parse_slots_str(row[c_unv]) if c_unv else []
            
            teachers[t_id] = Teacher(
                id=t_id,
                name=doc_val,
                cdc=cdc_val,
                is_part_time=is_pt,
                contract_hours=c_hours if is_pt else 18,
                max_working_days=m_days if is_pt else None,
                free_days=f_days,
                free_day_1=f_days[0] if f_days else None,
                free_day_2=f_days[1] if len(f_days) > 1 else None,
                prefer_late_entry=late_val,
                prefer_early_exit=early_val,
                max_daily_hours=mdh_val or 5,
                max_gap_hours=mgh_val if mgh_val is not None else 2,
                soft_avoid_slots=soft_slots,
                unavailable_slots=unavail_slots
            )
        else:
            if cdc_val and not teachers[t_id].cdc:
                teachers[t_id].cdc = cdc_val
            if c_sft and pd.notna(row[c_sft]):
                for s in _parse_slots_str(row[c_sft]):
                    if s not in teachers[t_id].soft_avoid_slots: teachers[t_id].soft_avoid_slots.append(s)
            if c_unv and pd.notna(row[c_unv]):
                for s in _parse_slots_str(row[c_unv]):
                    if s not in teachers[t_id].unavailable_slots: teachers[t_id].unavailable_slots.append(s)

        # 2. Parsing Classe
        c_id = _clean_id(cls_val)
        if c_id not in classes:
            grade_m = re.search(r'(\d+)', cls_val)
            grade_val = int(grade_m.group(1)) if grade_m else 1
            sec_m = re.search(r'([a-zA-Z])$', cls_val.strip())
            sec_val = sec_m.group(1).upper() if sec_m else "A"
            
            classes[c_id] = SchoolClass(
                id=c_id,
                name=cls_val,
                grade=grade_val,
                section=sec_val
            )

        # 3. Parsing Materia
        s_id = _clean_id(sub_val)
        if s_id not in subjects:
            color = DEFAULT_SUBJECT_COLORS.get(s_id, DEFAULT_SUBJECT_COLORS.get(sub_val.lower(), "#3498db"))
            subjects[s_id] = Subject(
                id=s_id,
                name=sub_val,
                color=color,
                cdc=cdc_val
            )

        # 4. Parsing Assegnazione Cattedra
        hours_val = _parse_int(row[c_hrs], default=2) if c_hrs else 2
        if hours_val <= 0: hours_val = 2
            
        double_val = _parse_bool(row[c_dbl]) if c_dbl else False
        max_daily_m = _parse_int(row[c_mda], default=2) if c_mda else 2
        
        assign_id = f"a_{c_id}_{s_id}_{t_id}_{len(assignments)}".lower()
        assignments.append(TeachingAssignment(
            id=assign_id,
            teacher_id=t_id,
            class_id=c_id,
            subject_id=s_id,
            hours_per_week=hours_val,
            force_double_hours=double_val,
            max_daily_hours=max_daily_m
        ))

    tot_ore_caricate = sum(a.hours_per_week for a in assignments)
    logs.append(f"Importati con successo: **{len(teachers)} docenti**, **{len(classes)} classi**, **{len(subjects)} materie**, **{len(assignments)} cattedre** (*Totale: **{tot_ore_caricate} ore settimanali***).")
    
    return TimetableProblem(
        config=config,
        teachers=teachers,
        classes=classes,
        subjects=subjects,
        rooms={},
        assignments=assignments
    ), logs


def parse_csv_timetable(file_content: str, base_config: Optional[SchoolConfig] = None) -> Tuple[TimetableProblem, List[str]]:
    """Legge una stringa CSV (auto-rilevando il separatore) e la converte in TimetableProblem."""
    first_line = file_content.split('\n')[0] if file_content else ""
    delimiter = ';' if ';' in first_line else (',' if ',' in first_line else '\t')
    df = pd.read_csv(io.StringIO(file_content), sep=delimiter, dtype=str)
    return parse_timetable_dataframe(df, base_config)


def parse_excel_timetable(file_bytes: bytes, base_config: Optional[SchoolConfig] = None) -> Tuple[TimetableProblem, List[str]]:
    """Legge un file Excel .xlsx in byte e lo converte in TimetableProblem."""
    df = pd.read_excel(io.BytesIO(file_bytes), dtype=str)
    return parse_timetable_dataframe(df, base_config)


# =============================================================
# MODULO RACCOLTA DESIDERATA DOCENTI (SELF-SERVICE / QUESTIONARIO)
# =============================================================
DESIDERATA_HEADERS = [
    "Docente",
    "CdC",
    "Part_Time",
    "Ore_Contratto",
    "Max_Giorni_Presenza",
    "Giorno_Libero_1",
    "Giorno_Libero_2",
    "Giorni_Entra_Tardi",
    "Giorni_Esce_Presto",
    "Max_Ore_Giorno",
    "Max_Ore_Buche",
    "Slot_Sconsigliati",
    "Slot_Indisponibili",
    "Note_Docente"
]

def generate_teacher_desiderata_form(problem: Optional[TimetableProblem] = None) -> bytes:
    """
    Genera un file Excel (.xlsx) formattato specificamente come 'Modulo Raccolta Desiderata Docenti'
    da inviare ai docenti o condividere su Google Drive / Microsoft Forms.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Desiderata Docenti"

    # Intestazione Modulo Istituzionale
    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    school_name = problem.config.school_name if problem else "Scuola Secondaria di I Grado"
    title_cell.value = f"📝 MODULO RACCOLTA DESIDERATA PERSONALI DOCENTI — {school_name.upper()}"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:N2")
    sub_cell = ws["A2"]
    sub_cell.value = "Istruzioni per il docente: Indicare giorni liberi (1ª/2ª scelta), giorni con preferenza entrata posticipata (No 1ª ora) o uscita anticipata (No ult. ora), e slot vietati per L.104/COE."
    sub_cell.font = Font(name="Calibri", size=10, italic=True, color="333333")
    sub_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # Header Colonne
    header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )

    for col_idx, h_text in enumerate(DESIDERATA_HEADERS, 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.value = h_text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[3].height = 28

    # Dati Docenti (o righe vuote se non presenti)
    row_num = 4
    if problem and problem.teachers:
        for t_id, t in problem.teachers.items():
            is_pt = "Si" if getattr(t, "is_part_time", False) else "No"
            ch = getattr(t, "contract_hours", 18) or 18
            max_d = getattr(t, "max_working_days", 5) if is_pt else 5
            
            f_days = getattr(t, "free_days", [])
            fd1 = f_days[0] if len(f_days) > 0 else (getattr(t, "free_day_1", "") or "")
            fd2 = f_days[1] if len(f_days) > 1 else (getattr(t, "free_day_2", "") or "")
            
            l_days = getattr(t, "late_entry_days", [])
            late_str = ", ".join(l_days) if l_days else ("Si" if getattr(t, "prefer_late_entry", False) else "No")
            
            e_days = getattr(t, "early_exit_days", [])
            early_str = ", ".join(e_days) if e_days else ("Si" if getattr(t, "prefer_early_exit", False) else "No")
            
            m_day = getattr(t, "max_daily_hours", 5) or 5
            m_gap = getattr(t, "max_gap_hours", 2) or 2
            
            # Format slots
            soft_str = ""
            if hasattr(t, "soft_avoid_slots") and t.soft_avoid_slots:
                parts = []
                for s in t.soft_avoid_slots:
                    if len(s) == 2 and s[0] < len(DAYS_OF_WEEK):
                        parts.append(f"{DAYS_OF_WEEK[s[0]]} {s[1] + 1}")
                soft_str = "; ".join(parts)
                
            unav_str = ""
            if hasattr(t, "unavailable_slots") and t.unavailable_slots:
                parts = []
                for s in t.unavailable_slots:
                    if len(s) == 2 and s[0] < len(DAYS_OF_WEEK):
                        parts.append(f"{DAYS_OF_WEEK[s[0]]} {s[1] + 1}")
                unav_str = "; ".join(parts)

            row_values = [
                t.name, getattr(t, "cdc", ""), is_pt, ch, max_d,
                fd1, fd2, late_str, early_str, m_day, m_gap, soft_str, unav_str, ""
            ]
            
            for col_idx, val in enumerate(row_values, 1):
                c = ws.cell(row=row_num, column=col_idx, value=val)
                c.alignment = Alignment(horizontal="center" if col_idx not in [1, 12, 13, 14] else "left", vertical="center")
                c.border = thin_border
            ws.row_dimensions[row_num].height = 20
            row_num += 1
    else:
        # 10 righe vuote d'esempio
        example_teachers = [
            ("Prof.ssa Bianchi", "A-22", "No", 18, 5, "Mercoledì", "", "Lunedì", "Venerdì", 5, 2, "Lunedì 1", "", "Entrata posticipata lunedì, uscita anticipata venerdì"),
            ("Prof. Romano", "A-28", "No", 18, 5, "Lunedì", "Venerdì", "Giovedì", "No", 4, 2, "", "Mercoledì 6", "Indisponibilità mercoledì 6a ora"),
            ("Prof.ssa Rossi", "A-24", "Si", 9, 3, "Lunedì", "Giovedì", "No", "No", 4, 1, "", "", "Part-time su 3 giorni"),
        ]
        for ex in example_teachers:
            row_values = list(ex) + [""]
            for col_idx, val in enumerate(row_values, 1):
                c = ws.cell(row=row_num, column=col_idx, value=val)
                c.alignment = Alignment(horizontal="center" if col_idx not in [1, 12, 13, 14] else "left", vertical="center")
                c.border = thin_border
            ws.row_dimensions[row_num].height = 20
            row_num += 1

    # Larghezza Colonne Ottimale
    col_widths = {
        "A": 26, "B": 12, "C": 12, "D": 14, "E": 20,
        "F": 18, "G": 18, "H": 14, "I": 14, "J": 15,
        "K": 15, "L": 24, "M": 24, "N": 35
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    out_io = io.BytesIO()
    wb.save(out_io)
    return out_io.getvalue()


def merge_teacher_desiderata_file(file_bytes_or_str: Any, problem: TimetableProblem, filename: str = "") -> Tuple[int, List[str]]:
    """
    Importa o fonde i desiderata compilati dai docenti (da file Excel o CSV)
    all'interno del problema corrente senza intaccare le classi o le cattedre esistenti.
    """
    logs = []
    if isinstance(file_bytes_or_str, bytes):
        if filename.lower().endswith(".csv"):
            content_str = file_bytes_or_str.decode('utf-8-sig', errors='replace')
            first_line = content_str.split('\n')[0] if content_str else ""
            delim = ';' if ';' in first_line else (',' if ',' in first_line else '\t')
            df = pd.read_csv(io.StringIO(content_str), sep=delim, dtype=str)
        else:
            # Salta le prime 2 righe se è il template formattato con titolo
            try:
                df = pd.read_excel(io.BytesIO(file_bytes_or_str), skiprows=2, dtype=str)
                # Se non ha le colonne giuste, prova a leggerlo senza skiprows
                if not any(c in str(df.columns) for c in ["Docente", "docente", "Nome"]):
                    df = pd.read_excel(io.BytesIO(file_bytes_or_str), dtype=str)
            except Exception:
                df = pd.read_excel(io.BytesIO(file_bytes_or_str), dtype=str)
    else:
        first_line = str(file_bytes_or_str).split('\n')[0] if file_bytes_or_str else ""
        delim = ';' if ';' in first_line else (',' if ',' in first_line else '\t')
        df = pd.read_csv(io.StringIO(str(file_bytes_or_str)), sep=delim, dtype=str)

    # Identifica le colonne nel dataframe caricato
    col_map = {}
    for c in df.columns:
        norm = str(c).lower().strip().replace(" ", "_").replace("-", "_")
        col_map[norm] = c

    def get_c(*keys):
        for k in keys:
            if k in col_map:
                return col_map[k]
        return None

    c_doc = get_c("docente", "nome_docente", "insegnante", "nome", "prof")
    if not c_doc:
        raise ValueError("Colonna 'Docente' non trovata nel file caricato.")

    c_cdc = get_c("cdc", "classe_di_concorso", "materia_cdc")
    c_pt = get_c("part_time", "pt", "tipo_contratto", "tempo_parziale")
    c_ch = get_c("ore_contratto", "ore", "monte_ore", "ore_settimanali")
    c_mwd = get_c("max_giorni_presenza", "max_giorni", "giorni_presenza", "giorni_max")
    c_fd1 = get_c("giorno_libero_1", "giorno_libero", "libero_1", "libero")
    c_fd2 = get_c("giorno_libero_2", "libero_2")
    c_late = get_c("giorni_entra_tardi", "entra_tardi", "no_1_ora", "no_1a_ora", "entrata_posticipata", "tardi")
    c_early = get_c("giorni_esce_presto", "esce_presto", "no_ultima_ora", "no_ult_ora", "uscita_anticipata", "presto")
    c_md = get_c("max_ore_giorno", "max_ore_giorno_docente", "max_ore_giornaliere")
    c_gap = get_c("max_ore_buche", "max_buche", "ore_buche")
    c_soft = get_c("slot_sconsigliati", "ore_sconsigliate", "sconsigliati")
    c_unav = get_c("slot_indisponibili", "indisponibilita", "indisponibili", "indisponibilita_assolute")

    num_updated = 0
    for _, row in df.iterrows():
        t_raw = str(row[c_doc]).strip()
        if not t_raw or t_raw.lower() in ["nan", "none", "docente", ""]:
            continue

        # Cerca il docente nel problema (per nome esatto o ID)
        target_teacher = None
        for t_obj in problem.teachers.values():
            if t_obj.name.lower() == t_raw.lower() or t_obj.id.lower() == t_raw.lower():
                target_teacher = t_obj
                break
        
        # Se non esiste, crealo
        if not target_teacher:
            new_id = f"doc_{t_raw.lower().replace(' ', '_').replace('.', '')}"
            target_teacher = Teacher(id=new_id, name=t_raw)
            problem.teachers[new_id] = target_teacher
            logs.append(f"Aggiunto nuovo docente: **{t_raw}**")

        # Aggiorna i desiderata
        if c_cdc and pd.notna(row[c_cdc]):
            target_teacher.cdc = str(row[c_cdc]).strip()

        if c_pt and pd.notna(row[c_pt]):
            target_teacher.is_part_time = _parse_bool(row[c_pt])

        if c_ch and pd.notna(row[c_ch]):
            target_teacher.contract_hours = _parse_int(row[c_ch], default=18)

        if c_mwd and pd.notna(row[c_mwd]):
            target_teacher.max_working_days = _parse_int(row[c_mwd], default=5)

        # Giorni liberi
        f_days = []
        if c_fds and pd.notna(row[c_fds]):
            f_days = _parse_free_days(row[c_fds])
        if c_fd1 and pd.notna(row[c_fd1]) and str(row[c_fd1]).strip() not in ["", "nan", "Nessuno"]:
            d1 = str(row[c_fd1]).strip()
            if d1 not in f_days: f_days.append(d1)
        if c_fd2 and pd.notna(row[c_fd2]) and str(row[c_fd2]).strip() not in ["", "nan", "Nessuno"]:
            d2 = str(row[c_fd2]).strip()
            if d2 not in f_days: f_days.append(d2)

        target_teacher.free_days = f_days
        target_teacher.free_day_1 = f_days[0] if len(f_days) > 0 else None
        target_teacher.free_day_2 = f_days[1] if len(f_days) > 1 else None

        if c_late and pd.notna(row[c_late]):
            l_val = str(row[c_late]).strip()
            l_days = _parse_free_days(l_val)
            target_teacher.late_entry_days = l_days
            target_teacher.prefer_late_entry = len(l_days) > 0 or _parse_bool(l_val)

        if c_early and pd.notna(row[c_early]):
            e_val = str(row[c_early]).strip()
            e_days = _parse_free_days(e_val)
            target_teacher.early_exit_days = e_days
            target_teacher.prefer_early_exit = len(e_days) > 0 or _parse_bool(e_val)

        if c_md and pd.notna(row[c_md]):
            target_teacher.max_daily_hours = _parse_int(row[c_md], default=5)

        if c_gap and pd.notna(row[c_gap]):
            target_teacher.max_gap_hours = _parse_int(row[c_gap], default=2)

        if c_soft and pd.notna(row[c_soft]):
            target_teacher.soft_avoid_slots = _parse_slots(row[c_soft])

        if c_unav and pd.notna(row[c_unav]):
            target_teacher.unavailable_slots = _parse_slots(row[c_unav])

        num_updated += 1

    logs.append(f"✅ Aggiornati i desiderata personali di **{num_updated} docenti** con successo!")
    return num_updated, logs


# =============================================================================
# FILE EXCEL MULTI-FOGLIO UNIFICATO SCUOLA (MASTER WORKBOOK)
# =============================================================================

def generate_unified_school_excel(problem: Optional[TimetableProblem] = None) -> bytes:
    """
    Genera un file Excel multi-foglio unificato (.xlsx) contenente l'intera banca dati
    della scuola:
      - 1. Struttura_e_Parametri
      - 2. Docenti
      - 3. Classi_e_Aule
      - 4. Cattedre_Curricolari
      - 5. Sostegno_e_DVA
      - 6. Classi_Aperte
      - 7. Guida_Compilazione
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # -------------------------------------------------------------
        # FOGLIO 1: Struttura_e_Parametri
        # -------------------------------------------------------------
        p_cfg = problem.config if problem else SchoolConfig()
        d_hours_str = ", ".join(str(h) for h in getattr(p_cfg, "daily_hours", [6]*p_cfg.num_days))
        struct_data = [
            ["Nome Scuola", p_cfg.school_name, "Nome dell'Istituto Scolastico"],
            ["Numero Giorni Settimanali", p_cfg.num_days, "5 (Settimana Corta) oppure 6 (Settimana Lunga)"],
            ["Ore Giornaliere", d_hours_str, "Ore per giorno separate da virgola (es. '6, 6, 6, 6, 6')"],
            ["Modello DADA", "Si" if getattr(p_cfg, "is_dada", True) else "No", "Si = Aule tematiche DADA; No = Modello tradizionale aule fisse"],
            ["Seconda Lingua Comunitaria", getattr(p_cfg, "second_language", "Spagnolo"), "Spagnolo, Francese, Tedesco o Personalizzata"],
            ["Blocchi Rigidi DADA (1-2, 3-4, 5-6)", "Si" if getattr(p_cfg, "is_strict_dada_slots", False) else "No", "Forza spostamenti aule solo agli intervalli"],
            ["Versione Formato File", "2.0", "Non modificare questo valore"]
        ]
        df_struct = pd.DataFrame(struct_data, columns=["Parametro", "Valore", "Descrizione"])
        df_struct.to_excel(writer, sheet_name="1_Struttura_e_Parametri", index=False)

        # -------------------------------------------------------------
        # FOGLIO 2: Docenti
        # -------------------------------------------------------------
        doc_rows = []
        if problem and problem.teachers:
            for t in sorted(problem.teachers.values(), key=lambda x: x.name):
                # Giorni liberi
                fl = getattr(t, "free_days", []) or []
                if not fl:
                    if getattr(t, "free_day_1", None): fl.append(t.free_day_1)
                    if getattr(t, "free_day_2", None): fl.append(t.free_day_2)
                fl_str = ", ".join(fl) if fl else ""

                # Slot sconsigliati
                soft_strs = []
                for d_i, h_i in getattr(t, "soft_avoid_slots", []):
                    if d_i < len(DAYS_OF_WEEK): soft_strs.append(f"{DAYS_OF_WEEK[d_i]} {h_i+1}")
                soft_s = ", ".join(soft_strs)

                # Slot indisponibili
                unav_strs = []
                for d_i, h_i in getattr(t, "unavailable_slots", []):
                    if d_i < len(DAYS_OF_WEEK): unav_strs.append(f"{DAYS_OF_WEEK[d_i]} {h_i+1}")
                unav_s = ", ".join(unav_strs)

                # Mappa CdC in materia leggibile in italiano
                cdc_raw = getattr(t, "cdc", "")
                materia_label = cdc_raw
                sec_lang = getattr(problem.config, "second_language", "Francese")
                cdc_map = {
                    "A-22": "Lettere (Italiano, Storia, Geografia)",
                    "A-28": "Matematica e Scienze",
                    "A-60": "Tecnologia",
                    "A-24": f"Lingue (Inglese / {sec_lang})",
                    "A-01": "Arte e Immagine",
                    "A-30": "Musica",
                    "A-48": "Scienze Motorie (Ginnastica)",
                    "Religione": "Religione Cattolica",
                    "ADMM": "Sostegno Didattico"
                }
                for c_k, c_v in cdc_map.items():
                    if c_k in cdc_raw:
                        materia_label = c_v
                        break

                doc_rows.append([
                    t.name,
                    materia_label,
                    "Si" if getattr(t, "is_part_time", False) else "No",
                    getattr(t, "contract_hours", 18),
                    getattr(t, "max_working_days", 5),
                    fl_str,
                    "Si" if getattr(t, "prefer_late_entry", False) else "No",
                    "Si" if getattr(t, "prefer_early_exit", False) else "No",
                    getattr(t, "max_daily_hours", 5),
                    getattr(t, "max_gap_hours", 2),
                    soft_s,
                    unav_s
                ])
        else:
            doc_rows = [
                ["Prof.ssa Bianchi M.", "Lettere (Italiano, Storia, Geografia)", "No", 18, 5, "", "No", "Si", 5, 2, "", ""],
                ["Prof. Verdi G.", "Matematica e Scienze", "No", 18, 5, "", "Si", "No", 5, 1, "", ""],
                ["Prof.ssa Colombo S.", "Francese", "Si", 12, 3, "Lunedì, Mercoledì", "No", "No", 4, 1, "Mercoledì 1", "Lunedì 6"]
            ]
        doc_cols = [
            "Docente", "Materia_Insegnamento", "Part_Time", "Ore_Contratto", "Max_Giorni_Presenza",
            "Giorni_Liberi", "Entra_Tardi", "Esce_Presto", "Max_Ore_Giorno", "Max_Ore_Buche",
            "Slot_Sconsigliati", "Slot_Indisponibili"
        ]
        df_doc = pd.DataFrame(doc_rows, columns=doc_cols)
        df_doc.to_excel(writer, sheet_name="2_Docenti", index=False)

        # -------------------------------------------------------------
        # FOGLIO 3: Classi
        # -------------------------------------------------------------
        class_rows = []
        if problem and problem.classes:
            for c in sorted(problem.classes.values(), key=lambda x: (x.grade, x.section)):
                class_rows.append([c.name, c.grade, c.section])
        else:
            class_rows = [
                ["1ª A", 1, "A"],
                ["2ª A", 2, "A"],
                ["3ª A", 3, "A"]
            ]
        class_cols = ["Classe", "Anno", "Sezione"]
        df_class = pd.DataFrame(class_rows, columns=class_cols)
        df_class.to_excel(writer, sheet_name="3_Classi", index=False)

        # -------------------------------------------------------------
        # FOGLIO 4: Aule_e_Laboratori
        # -------------------------------------------------------------
        room_rows = []
        if problem and problem.rooms:
            for r in sorted(problem.rooms.values(), key=lambda x: x.name):
                subs_str = ", ".join(r.subject_ids) if getattr(r, "subject_ids", None) else ""
                t_names = []
                for tid in getattr(r, "teacher_ids", []):
                    if tid in problem.teachers:
                        t_names.append(problem.teachers[tid].name)
                    else:
                        t_names.append(tid)
                t_str = ", ".join(t_names) if t_names else ""
                room_rows.append([
                    r.name,
                    subs_str,
                    r.capacity or 1,
                    r.priority or 1,
                    "Si" if getattr(r, "is_special_lab", False) else "No",
                    t_str
                ])
        else:
            room_rows = [
                ["Aula 101 - Lettere A", "ita, sto, geo", 1, 1, "No", "Prof. Valenti S., Prof.ssa Montanari G."],
                ["Aula 102 - Matematica A", "mat, sci", 1, 1, "No", "Prof. Marchetti E."],
                ["Palestra 1", "mot", 2, 1, "Si", "Prof.ssa Rossetti M., Prof.ssa Leone P."],
                ["Laboratorio Scienze", "sci", 1, 1, "Si", "Prof. Serra G."]
            ]
        room_cols = ["Nome_Aula", "Materie_Assegnate", "Capienza_Classi", "Priorita", "Laboratorio_Speciale", "Docenti_Assegnati"]
        df_room = pd.DataFrame(room_rows, columns=room_cols)
        df_room.to_excel(writer, sheet_name="4_Aule_e_Laboratori", index=False)

        # -------------------------------------------------------------
        # FOGLIO 5: Cattedre_Curricolari
        # -------------------------------------------------------------
        assign_rows = []
        if problem and problem.assignments:
            sorted_a = sorted(
                problem.assignments,
                key=lambda a: (
                    problem.teachers.get(a.teacher_id).name if a.teacher_id in problem.teachers else "",
                    problem.classes.get(a.class_id).name if a.class_id in problem.classes else ""
                )
            )
            for a in sorted_a:
                t = problem.teachers.get(a.teacher_id)
                c = problem.classes.get(a.class_id)
                s = problem.subjects.get(a.subject_id)
                assign_rows.append([
                    t.name if t else a.teacher_id,
                    c.name if c else a.class_id,
                    s.name if s else a.subject_id,
                    getattr(t, "cdc", "") if t else (getattr(s, "cdc", "") if s else ""),
                    a.hours_per_week,
                    "Si" if a.force_double_hours else "No",
                    a.max_daily_hours or 2
                ])
        else:
            assign_rows = [
                ["Prof.ssa Bianchi M.", "1ª A", "Italiano", "A-22", 6, "Si", 2],
                ["Prof.ssa Bianchi M.", "1ª A", "Storia", "A-22", 2, "No", 2],
                ["Prof. Verdi G.", "1ª A", "Matematica", "A-28", 4, "Si", 2],
                ["Prof. Verdi G.", "1ª A", "Scienze", "A-28", 2, "No", 2],
                ["Prof.ssa Colombo S.", "1ª A", "Seconda Lingua (Spagnolo)", "A-24", 2, "No", 2]
            ]
        assign_cols = ["Docente", "Classe", "Materia", "CdC", "Ore_Settimanali", "Ore_Doppie", "Max_Ore_Giorno_Materia"]
        df_assign = pd.DataFrame(assign_rows, columns=assign_cols)
        df_assign.to_excel(writer, sheet_name="5_Cattedre_Curricolari", index=False)

        # -------------------------------------------------------------
        # FOGLIO 5: Sostegno_e_DVA
        # -------------------------------------------------------------
        sost_rows = []
        if problem and getattr(problem, "students_dva", None):
            for st_id, st_obj in sorted(problem.students_dva.items(), key=lambda x: x[1].name):
                c_name = problem.classes.get(st_obj.class_id).name if st_obj.class_id in problem.classes else st_obj.class_id
                
                # Dati alunno DVA
                is_grave_str = "Si" if getattr(st_obj, "is_severe_coverage", False) else "No"
                
                # Materie preferite da coprire per l'alunno (converte id in nomi leggibili es. ita, mat -> Italiano, Matematica)
                sub_names = []
                for sid in getattr(st_obj, "preferred_subjects", []):
                    if sid in problem.subjects:
                        sub_names.append(problem.subjects[sid].name)
                    else:
                        sub_names.append(sid)
                mat_coperte_str = ", ".join(sub_names) if sub_names else "Tutte le discipline"

                # Assegnazioni docenti per questo studente
                sa_list = [sa for sa in getattr(problem, "support_assignments", []) if getattr(sa, "student_id", getattr(sa, "student_dva_id", None)) == st_id]
                if sa_list:
                    for sa in sa_list:
                        t = problem.teachers.get(sa.teacher_id)
                        t_name = t.name if t else sa.teacher_id
                        pref_areas = ", ".join(getattr(t, "preferred_areas", [])) if t else ""
                        
                        # Parametri contrattuali e desiderata completi docente sostegno
                        is_pt = "Si" if getattr(t, "is_part_time", False) else "No"
                        ch = getattr(t, "contract_hours", 18) or 18
                        max_d = getattr(t, "max_working_days", 5) if is_pt == "Si" else 5
                        m_day = getattr(t, "max_daily_hours", 5) or 5
                        m_gap = getattr(t, "max_gap_hours", 2) if t else 2
                        
                        fl = getattr(t, "free_days", []) or []
                        fl_str = ", ".join(fl) if fl else ""
                        late_str = "Si" if (t and getattr(t, "prefer_late_entry", False)) else "No"
                        early_str = "Si" if (t and getattr(t, "prefer_early_exit", False)) else "No"
                        
                        unav_strs = []
                        if t and getattr(t, "unavailable_slots", []):
                            for d_i, h_i in t.unavailable_slots:
                                if d_i < len(DAYS_OF_WEEK): unav_strs.append(f"{DAYS_OF_WEEK[d_i]} {h_i+1}")
                        unav_s = ", ".join(unav_strs)

                        sost_rows.append([
                            st_obj.name,
                            c_name,
                            st_obj.weekly_hours,
                            is_grave_str,
                            mat_coperte_str,
                            t_name,
                            sa.hours_per_week,
                            pref_areas or "Tutte",
                            is_pt,
                            ch,
                            max_d,
                            m_day,
                            m_gap,
                            fl_str,
                            late_str,
                            early_str,
                            unav_s
                        ])
                else:
                    sost_rows.append([
                        st_obj.name,
                        c_name,
                        st_obj.weekly_hours,
                        is_grave_str,
                        mat_coperte_str,
                        "",
                        0,
                        "Tutte",
                        "No",
                        18,
                        5,
                        5,
                        2,
                        "",
                        "No",
                        "No",
                        ""
                    ])
        else:
            sost_rows = [
                ["Alunno Rossi M. (1ª A)", "1ª A", 9, "No", "Matematica, Scienze", "Prof. Gentile (Sostegno 18h)", 9, "scientifica", "No", 18, 5, 5, 2, "", "No", "No", ""],
                ["Alunno Bianchi F. (2ª A)", "2ª A", 18, "Si", "Italiano, Storia, Matematica", "Prof. Marini (Sostegno 18h)", 18, "umanistica, scientifica", "Si", 12, 3, 4, 1, "Mercoledì", "Si", "No", "Venerdì 6"]
            ]
        sost_cols = [
            "Studente_DVA", "Classe", "Ore_Totali_Richieste", "Gravita_Caso_Grave", "Materie_Da_Coprire",
            "Docente_Sostegno", "Ore_Assegnate", "Aree_Disciplinari_Preferite", "Part_Time", "Ore_Contratto", 
            "Max_Giorni_Presenza", "Max_Ore_Giorno", "Max_Ore_Buche", "Giorni_Liberi_Docente", "Entra_Tardi", "Esce_Presto", "Slot_Indisponibili"
        ]
        df_sost = pd.DataFrame(sost_rows, columns=sost_cols)
        df_sost.to_excel(writer, sheet_name="5_Sostegno_e_DVA", index=False)

        # -------------------------------------------------------------
        # FOGLIO 6: Classi_Aperte_Parallelismi
        # -------------------------------------------------------------
        par_rows = []
        if problem and getattr(problem, "parallel_groups", None):
            for pg in problem.parallel_groups:
                c_names = []
                for cid in pg.class_ids:
                    c_obj = problem.classes.get(cid)
                    c_names.append(c_obj.name if c_obj else cid)
                s_obj = problem.subjects.get(pg.subject_id)
                s_name = s_obj.name if s_obj else pg.subject_id
                par_rows.append([
                    pg.name,
                    s_name,
                    ", ".join(c_names),
                    getattr(pg, "parallel_hours", getattr(pg, "hours_per_week", 2))
                ])
        else:
            par_rows = [
                ["Parallelismo Seconde - Spagnolo/Francese", "Seconda Lingua", "2ª A, 2ª B", 2]
            ]
        par_cols = ["Nome_Gruppo", "Materia", "Classi_Coinvolte", "Ore_Settimanali"]
        df_par = pd.DataFrame(par_rows, columns=par_cols)
        df_par.to_excel(writer, sheet_name="6_Classi_Aperte", index=False)

        # -------------------------------------------------------------
        # FOGLIO 7: Guida_Compilazione
        # -------------------------------------------------------------
        guide_rows = [
            ["Foglio", "Descrizione e Regole di Compilazione"],
            ["1_Struttura_e_Parametri", "Configura il nome della scuola, se lavori a 5 o 6 giorni, le ore per giorno e il modello DADA."],
            ["2_Docenti", "Elenco di tutti i docenti. Per i Part-Time indica 'Si' e specifica max giorni di presenza e ore contratto."],
            ["3_Classi_e_Aule", "Elenco delle classi e dell'aula DADA / laboratorio assegnata come base."],
            ["4_Cattedre_Curricolari", "Abbinamento Docente-Classe-Materia con monte ore settimanale e preferenza ore doppie consecutive."],
            ["5_Sostegno_e_DVA", "Elenco studenti con certificazione DVA e docenti di sostegno assegnati con relativo monte ore."],
            ["6_Classi_Aperte", "Opzionale: definisce materie svolte contemporaneamente a classi aperte (es. opzioni lingue)."],
            ["Consiglio Operativo", "Puoi modificare i dati su Excel e ricaricare il file su Orario Scolastico Facile con 1 clic per un ripristino totale!"]
        ]
        df_guide = pd.DataFrame(guide_rows[1:], columns=guide_rows[0])
        df_guide.to_excel(writer, sheet_name="7_Guida_Compilazione", index=False)

        # -------------------------------------------------------------
        # STYLING GRAFICO OPENPYXL PER TUTTI I FOGLI
        # -------------------------------------------------------------
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Blu Navy Elegante
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        for sheetname in writer.sheets:
            ws = writer.sheets[sheetname]
            max_cols = ws.max_column
            max_rows = ws.max_row

            for col_idx in range(1, max_cols + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align

                # Larghezza colonna automatica
                vals = [str(ws.cell(row=r, column=col_idx).value or "") for r in range(1, max_rows + 1)]
                max_w = max(len(v) for v in vals) if vals else 10
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = max(max_w + 4, 16)

            for r in range(2, max_rows + 1):
                for c in range(1, max_cols + 1):
                    cell = ws.cell(row=r, column=c)
                    cell.border = thin_border

    return output.getvalue()


def parse_unified_school_excel(file_bytes: bytes, base_config: Optional[SchoolConfig] = None) -> Tuple[TimetableProblem, List[str]]:
    """
    Parser intelligente che supporta sia il nuovo formato multi-foglio unificato (.xlsx)
    sia i file legacy a foglio singolo.
    """
    logs = []
    xl = pd.ExcelFile(io.BytesIO(file_bytes))
    sheet_names = xl.sheet_names

    # Controlla se è il nuovo formato multi-foglio
    has_struct = any("struttura" in s.lower() or "1_" in s for s in sheet_names)
    has_doc = any("docenti" in s.lower() or "2_" in s for s in sheet_names)
    has_class = any("classi" in s.lower() or "3_" in s for s in sheet_names)
    has_assign = any("cattedre" in s.lower() or "4_" in s for s in sheet_names)

    if not (has_struct or has_doc or has_assign):
        # Fallback su parser a foglio singolo
        df_single = xl.parse(sheet_names[0])
        return parse_timetable_dataframe(df_single, base_config)

    # 1. Parsing Struttura
    config = base_config or SchoolConfig()
    for s in sheet_names:
        if "struttura" in s.lower() or "1_" in s:
            df_s = xl.parse(s)
            if "Parametro" in df_s.columns and "Valore" in df_s.columns:
                for _, r in df_s.iterrows():
                    param = str(r["Parametro"]).strip().lower()
                    val = str(r["Valore"]).strip()
                    if "nome scuola" in param and val:
                        config.school_name = val
                    elif "giorni" in param:
                        config.num_days = _parse_int(val, default=5)
                    elif "ore giornaliere" in param and val:
                        try:
                            config.daily_hours = [int(x.strip()) for x in val.split(",") if x.strip()]
                        except:
                            config.daily_hours = [6] * config.num_days
                    elif "dada" in param:
                        config.is_dada = _parse_bool(val)
                    elif "lingua" in param and val:
                        config.second_language = val
                    elif "rigidi" in param:
                        config.is_strict_dada_slots = _parse_bool(val)
            logs.append(f"⚙️ Configurazione scuola caricata: **{config.school_name}** ({config.num_days} giorni, DADA: {'Sì' if config.is_dada else 'No'}).")
            break

    teachers: Dict[str, Teacher] = {}
    classes: Dict[str, SchoolClass] = {}
    classrooms: Dict[str, Classroom] = {}
    subjects: Dict[str, Subject] = {}
    assignments: List[TeachingAssignment] = []
    students_dva: Dict[str, StudentDVA] = {}
    support_assignments: List[SupportAssignment] = []
    parallel_groups: List[ParallelGroup] = []

    # 2. Parsing Docenti
    for s in sheet_names:
        if "docenti" in s.lower() or "2_" in s:
            df_d = xl.parse(s)
            for _, r in df_d.iterrows():
                doc_name = str(r.get("Docente", "")).strip()
                if not doc_name or doc_name.lower() in ["nan", "none", ""]:
                    continue
                t_id = "doc_" + _clean_id(doc_name)
                is_pt = _parse_bool(r.get("Part_Time", False))
                c_hours = _parse_int(r.get("Ore_Contratto", 18), default=(12 if is_pt else 18))
                m_days = _parse_int(r.get("Max_Giorni_Presenza", config.num_days), default=(3 if is_pt else config.num_days))
                
                f_days = _parse_free_days(r.get("Giorni_Liberi", ""))
                late_val = _parse_bool(r.get("Entra_Tardi", False))
                early_val = _parse_bool(r.get("Esce_Presto", False))
                mdh = _parse_int(r.get("Max_Ore_Giorno", 5), default=5)
                mgh = _parse_int(r.get("Max_Ore_Buche", 2), default=2)
                
                soft_slots = _parse_slots_str(r.get("Slot_Sconsigliati", ""))
                unav_slots = _parse_slots_str(r.get("Slot_Indisponibili", ""))

                raw_cdc = str(r.get("Materia_Insegnamento", r.get("CdC", r.get("Materia", "")))).strip()
                if raw_cdc.lower() == "nan": raw_cdc = ""
                # Converti descrizione in CdC pulita se necessario
                clean_cdc = raw_cdc
                if "lettere" in raw_cdc.lower() or "italiano" in raw_cdc.lower(): clean_cdc = "A-22"
                elif "matematica" in raw_cdc.lower() or "scienze" in raw_cdc.lower() and "motoria" not in raw_cdc.lower(): clean_cdc = "A-28"
                elif "tecnologia" in raw_cdc.lower() or "coding" in raw_cdc.lower(): clean_cdc = "A-60"
                elif "inglese" in raw_cdc.lower() or "francese" in raw_cdc.lower() or "spagnolo" in raw_cdc.lower() or "tedesco" in raw_cdc.lower() or "lingua" in raw_cdc.lower(): clean_cdc = "A-24"
                elif "arte" in raw_cdc.lower() or "immagine" in raw_cdc.lower(): clean_cdc = "A-01"
                elif "musica" in raw_cdc.lower(): clean_cdc = "A-30"
                elif "motoria" in raw_cdc.lower() or "ginnastica" in raw_cdc.lower() or "sport" in raw_cdc.lower(): clean_cdc = "A-48"
                elif "religione" in raw_cdc.lower(): clean_cdc = "Religione"
                elif "sostegno" in raw_cdc.lower(): clean_cdc = "ADMM"

                teachers[t_id] = Teacher(
                    id=t_id,
                    name=doc_name,
                    cdc=clean_cdc,
                    is_part_time=is_pt,
                    contract_hours=c_hours,
                    max_working_days=m_days if is_pt else None,
                    free_days=f_days,
                    free_day_1=f_days[0] if f_days else None,
                    free_day_2=f_days[1] if len(f_days) > 1 else None,
                    prefer_late_entry=late_val,
                    prefer_early_exit=early_val,
                    max_daily_hours=mdh,
                    max_gap_hours=mgh,
                    soft_avoid_slots=soft_slots,
                    unavailable_slots=unav_slots
                )
            logs.append(f"👥 Caricati **{len(teachers)} docenti**.")
            break

    # 3. Parsing Classi
    for s in sheet_names:
        if "classi" in s.lower() or "3_" in s:
            df_c = xl.parse(s)
            for _, r in df_c.iterrows():
                c_name = str(r.get("Classe", "")).strip()
                if not c_name or c_name.lower() in ["nan", "none", ""]:
                    continue
                c_id = _clean_id(c_name)
                grade_v = _parse_int(r.get("Anno", 1), default=1)
                sec_v = str(r.get("Sezione", "A")).strip()
                classes[c_id] = SchoolClass(id=c_id, name=c_name, grade=grade_v, section=sec_v)
            logs.append(f"🏫 Caricate **{len(classes)} classi**.")
            break

    # 4. Parsing Aule & Ambienti DADA
    for s in sheet_names:
        if "aule" in s.lower() or "4_" in s:
            df_r = xl.parse(s)
            for _, r in df_r.iterrows():
                r_name = str(r.get("Nome_Aula", "")).strip()
                if not r_name or r_name.lower() in ["nan", "none", ""]:
                    continue
                r_id = "room_" + _clean_id(r_name)
                subs_raw = str(r.get("Materie_Assegnate", "")).strip()
                sub_ids = [x.strip().lower() for x in subs_raw.split(",") if x.strip() and subs_raw.lower() != "nan"]
                r_cap = _parse_int(r.get("Capienza_Classi", r.get("Capienza_Aula", 1)), default=1)
                r_prio = _parse_int(r.get("Priorita", 1), default=1)
                r_spec = _parse_bool(r.get("Laboratorio_Speciale", False))
                docs_raw = str(r.get("Docenti_Assegnati", r.get("Docenti", ""))).strip()
                t_ids = []
                if docs_raw and docs_raw.lower() != "nan":
                    for d_entry in docs_raw.split(","):
                        d_clean = d_entry.strip()
                        if not d_clean:
                            continue
                        matched_tid = None
                        for cand_tid, cand_t in teachers.items():
                            if cand_t.name.lower() == d_clean.lower() or cand_tid.lower() == d_clean.lower():
                                matched_tid = cand_tid
                                break
                        if matched_tid:
                            t_ids.append(matched_tid)
                        else:
                            # id pulito
                            t_ids.append("doc_" + _clean_id(d_clean))

                classrooms[r_id] = Classroom(
                    id=r_id,
                    name=r_name,
                    subject_ids=sub_ids,
                    capacity=r_cap,
                    priority=r_prio,
                    is_special_lab=r_spec,
                    teacher_ids=t_ids
                )
            logs.append(f"🏛️ Caricati **{len(classrooms)} ambienti e aule DADA**.")
            break

    # 5. Parsing Cattedre Curricolari
    for s in sheet_names:
        if "cattedre" in s.lower() or "curricolari" in s.lower() or "5_" in s:
            df_a = xl.parse(s)
            for _, r in df_a.iterrows():
                t_name = str(r.get("Docente", "")).strip()
                c_name = str(r.get("Classe", "")).strip()
                s_name = str(r.get("Materia", "")).strip()
                if not t_name or not c_name or not s_name or t_name.lower() == "nan":
                    continue
                
                t_id = "doc_" + _clean_id(t_name)
                if t_id not in teachers:
                    teachers[t_id] = Teacher(id=t_id, name=t_name)
                    
                c_id = _clean_id(c_name)
                if c_id not in classes:
                    classes[c_id] = SchoolClass(id=c_id, name=c_name, grade=1, section="A")

                s_id = _clean_id(s_name)
                if s_id not in subjects:
                    cdc_v = str(r.get("CdC", "")).strip() if pd.notna(r.get("CdC", "")) else ""
                    subjects[s_id] = Subject(
                        id=s_id,
                        name=s_name,
                        color=DEFAULT_SUBJECT_COLORS.get(s_id, DEFAULT_SUBJECT_COLORS.get(s_name.lower(), "#3498db")),
                        cdc=cdc_v
                    )

                h_val = _parse_int(r.get("Ore_Settimanali", 2), default=2)
                dbl_val = _parse_bool(r.get("Ore_Doppie", False))
                max_d = _parse_int(r.get("Max_Ore_Giorno_Materia", 2), default=2)

                a_id = f"a_{c_id}_{s_id}_{t_id}_{len(assignments)}".lower()
                assignments.append(TeachingAssignment(
                    id=a_id,
                    teacher_id=t_id,
                    class_id=c_id,
                    subject_id=s_id,
                    hours_per_week=h_val,
                    force_double_hours=dbl_val,
                    max_daily_hours=max_d
                ))
            logs.append(f"📚 Caricate **{len(assignments)} cattedre curricolari**.")
            break

    # 6. Parsing Sostegno e DVA
    for s in sheet_names:
        if "sostegno" in s.lower() or "dva" in s.lower() or "6_" in s:
            df_sost = xl.parse(s)
            for _, r in df_sost.iterrows():
                dva_name = str(r.get("Studente_DVA", "")).strip()
                c_name = str(r.get("Classe", "")).strip()
                if not dva_name or not c_name or dva_name.lower() == "nan":
                    continue
                dva_id = "dva_" + _clean_id(dva_name)
                c_id = _clean_id(c_name)
                tot_hrs = _parse_int(r.get("Ore_Totali_Richieste", 9), default=9)

                # Gravità e Materie da coprire
                is_grave = _parse_bool(r.get("Gravita_Caso_Grave", r.get("Gravita", False)))
                
                pref_sub_ids = []
                raw_subs = str(r.get("Materie_Da_Coprire", r.get("Materie", ""))).strip()
                if raw_subs and raw_subs.lower() not in ["nan", "none", "tutte", "tutte le discipline", ""]:
                    for item in raw_subs.split(","):
                        item_clean = item.strip().lower()
                        if not item_clean: continue
                        # Cerca corrispondenza nelle materie
                        matched_s = None
                        for s_k, s_obj in subjects.items():
                            if s_obj.name.lower() == item_clean or s_k.lower() == item_clean:
                                matched_s = s_k
                                break
                        if matched_s:
                            pref_sub_ids.append(matched_s)
                        else:
                            # mapping euristico rapido
                            if "ita" in item_clean or "lettere" in item_clean: pref_sub_ids.append("ita")
                            elif "mat" in item_clean: pref_sub_ids.append("mat")
                            elif "sci" in item_clean: pref_sub_ids.append("sci")
                            elif "ing" in item_clean: pref_sub_ids.append("ing")
                            elif "sto" in item_clean: pref_sub_ids.append("sto")
                            elif "geo" in item_clean: pref_sub_ids.append("geo")
                            elif "tec" in item_clean: pref_sub_ids.append("tec")
                            elif "art" in item_clean: pref_sub_ids.append("art")
                            elif "mus" in item_clean: pref_sub_ids.append("mus")
                            elif "mot" in item_clean: pref_sub_ids.append("mot")

                if dva_id not in students_dva:
                    students_dva[dva_id] = StudentDVA(
                        id=dva_id,
                        name=dva_name,
                        class_id=c_id,
                        weekly_hours=tot_hrs,
                        is_severe_coverage=is_grave,
                        preferred_subjects=pref_sub_ids
                    )
                else:
                    if is_grave:
                        students_dva[dva_id].is_severe_coverage = True
                    if pref_sub_ids:
                        students_dva[dva_id].preferred_subjects = list(set(students_dva[dva_id].preferred_subjects + pref_sub_ids))

                doc_sost = str(r.get("Docente_Sostegno", "")).strip()
                assign_hrs = _parse_int(r.get("Ore_Assegnate", tot_hrs), default=tot_hrs)
                if doc_sost and doc_sost.lower() not in ["nan", "none", ""]:
                    t_id = "doc_" + _clean_id(doc_sost)
                    if t_id not in teachers:
                        teachers[t_id] = Teacher(id=t_id, name=doc_sost, cdc="ADMM Sostegno")
                    
                    pref_areas = []
                    if pd.notna(r.get("Aree_Disciplinari_Preferite", "")):
                        pref_areas = [x.strip().lower() for x in str(r.get("Aree_Disciplinari_Preferite")).split(",") if x.strip()]
                    teachers[t_id].preferred_areas = pref_areas or ["umanistica", "scientifica"]

                    # Carica parametri contrattuali e desiderata completi del docente di sostegno
                    if "Part_Time" in r and pd.notna(r["Part_Time"]):
                        is_pt_val = _parse_bool(r["Part_Time"])
                        teachers[t_id].is_part_time = is_pt_val
                    if "Ore_Contratto" in r and pd.notna(r["Ore_Contratto"]):
                        teachers[t_id].contract_hours = _parse_int(r["Ore_Contratto"], default=18)
                    if "Max_Giorni_Presenza" in r and pd.notna(r["Max_Giorni_Presenza"]):
                        teachers[t_id].max_working_days = _parse_int(r["Max_Giorni_Presenza"], default=config.num_days)
                    if "Max_Ore_Giorno" in r and pd.notna(r["Max_Ore_Giorno"]):
                        teachers[t_id].max_daily_hours = _parse_int(r["Max_Ore_Giorno"], default=5)
                    if "Giorni_Liberi_Docente" in r and pd.notna(r["Giorni_Liberi_Docente"]):
                        f_days = _parse_free_days(r["Giorni_Liberi_Docente"])
                        teachers[t_id].free_days = f_days
                        teachers[t_id].free_day_1 = f_days[0] if f_days else None
                        teachers[t_id].free_day_2 = f_days[1] if len(f_days) > 1 else None
                    if "Entra_Tardi" in r and pd.notna(r["Entra_Tardi"]):
                        teachers[t_id].prefer_late_entry = _parse_bool(r["Entra_Tardi"])
                    if "Esce_Presto" in r and pd.notna(r["Esce_Presto"]):
                        teachers[t_id].prefer_early_exit = _parse_bool(r["Esce_Presto"])
                    if "Max_Ore_Buche" in r and pd.notna(r["Max_Ore_Buche"]):
                        teachers[t_id].max_gap_hours = _parse_int(r["Max_Ore_Buche"], default=2)
                    if "Slot_Indisponibili" in r and pd.notna(r["Slot_Indisponibili"]):
                        teachers[t_id].unavailable_slots = _parse_slots_str(r["Slot_Indisponibili"])

                    sa_id = f"sa_{dva_id}_{t_id}_{len(support_assignments)}"
                    support_assignments.append(SupportAssignment(
                        id=sa_id,
                        teacher_id=t_id,
                        student_id=dva_id,
                        class_id=c_id,
                        hours_per_week=assign_hrs
                    ))
            if students_dva:
                logs.append(f"🤝 Caricati **{len(students_dva)} studenti DVA** e **{len(support_assignments)} abbinamenti sostegno** (con desiderata docenti).")
            break

    # 7. Parsing Classi Aperte / Parallelismi
    for s in sheet_names:
        if "parallel" in s.lower() or "classi_aperte" in s.lower() or "7_" in s:
            df_p = xl.parse(s)
            for _, r in df_p.iterrows():
                grp_name = str(r.get("Nome_Gruppo", "")).strip()
                sub_name = str(r.get("Materia", "")).strip()
                cls_str = str(r.get("Classi_Coinvolte", "")).strip()
                if not grp_name or not cls_str or grp_name.lower() == "nan":
                    continue
                grp_id = "pg_" + _clean_id(grp_name)
                s_id = _clean_id(sub_name) if sub_name else "seconda_lingua"
                
                cls_ids = []
                for c_item in re.split(r'[,;/|]+', cls_str):
                    c_cl = c_item.strip()
                    if c_cl:
                        cls_ids.append(_clean_id(c_cl))

                h_val = _parse_int(r.get("Ore_Settimanali", 2), default=2)
                parallel_groups.append(ParallelGroup(
                    id=grp_id,
                    name=grp_name,
                    subject_id=s_id,
                    class_ids=cls_ids,
                    parallel_hours=h_val
                ))
            if parallel_groups:
                logs.append(f"🔀 Caricati **{len(parallel_groups)} gruppi a classi aperte / parallelismi**.")
            break

    config.parallel_groups = parallel_groups
    prob = TimetableProblem(
        config=config,
        teachers=teachers,
        classes=classes,
        subjects=subjects,
        rooms=classrooms,
        assignments=assignments,
        students_dva=students_dva,
        support_assignments=support_assignments
    )

    logs.append(f"🎉 **Importazione Unificata Completata con successo!** Totale: **{len(teachers)} docenti**, **{len(classes)} classi**, **{len(assignments)} cattedre curricolari**.")
    return prob, logs


